from django.shortcuts import render
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import TemplateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .tasks import get_data_for_text,get_data_for_xlxs
from django.utils import timezone
from django.contrib import messages
import pandas as pd
from django.shortcuts import render , redirect 
from departments.models import College,Department
from django_celery_beat.models import PeriodicTask,IntervalSchedule
from django_celery_results.models import TaskResult
import openpyxl
import json
from UrlSearch.models import BasicInformation
from .models import TasksInfo
# from TrackYourAlumni.models import TasksInfo
#from . import function
#from . import function
# Create your views here.

#get file from user
def getProfile(request):
    if request.method=="POST":
        file=request.FILES["url_file"]

        print("uploaded")
    return render(request,'temp.html')
 
@method_decorator(login_required, name="dispatch")
class SearchPeopleFromUrl(TemplateView):
    template_name= "temp.html"
    success_url = reverse_lazy("urlsearch:getProfile")

    def get_context_data(self, **kwargs):
        context =  super().get_context_data(**kwargs)

        info={}
        all_task_info=TasksInfo.objects.filter(college_name=self.request.user.college_name)
       
        for task in all_task_info:
            task_result=TaskResult.objects.filter(task_id=task.task).first()
            if task_result:
                result=task_result.result
                if "Done" in result:
                    string_result=result.strip('\"')    

                else:
                    temp=json.loads(result)
                    if 'current' in temp:
                        print(temp,"---temp")
                        try:
                            string_result=str(temp["current"])+' out of '+str(temp["total"])+' ('+str(temp["percent"])+'%)'
                        except:
                            string_result="Pending"
                    elif "exc_type" in temp:
                        string_result=str(temp['exc_type']+" Execption Please reupload this file")            
                    # string_result=""
                    task_result.date_done=""
            else:
                string_result="Pendding"

            info[task.task]={"task_info":task,"task_result":task_result,"result":string_result}


        context["tasks"]=info
        return context

    def post(self, request, *args, **kwargs):
        user=request.user
        if request.user.is_role_admin or request.user.is_role_superuser:
            task_name=request.POST.get("task_name")
            file=request.FILES.get('profile_url_file')
            
            filestr = str(file)
            print("Uploaded ",file)
            if filestr.lower().endswith('.txt'):
                # task1=get_data_for_text.delay(file, user,department,record_year)
                # TakeTaskInfo(task1,department,record_year,user,file)
                # get_data_for_text(file,user)
                messages.error(
                        request,
                        f"Our Text File service temporery anvailable ",
                    )
                return redirect(
                        reverse_lazy("urlsearch:getProfile")
                    )
            
            elif filestr.lower().endswith('.xlsx'):
                read_file =  pd.read_excel(file)
                if read_file.empty:
                    messages.error(
                        request,
                        f"Failed upload: Selected file is empty. Please fill-up the valid data in the Excel File.   ",
                    )
                    return redirect(
                        reverse_lazy("urlsearch:getProfile")
                    )

                wb = openpyxl.load_workbook(file)
                worksheets = wb.sheetnames
                fields_dict = {}
                value_dict={}
                valid_columns_list = ['LINKEDIN PROFILE URL','STUDENT NAME','COLLEGE ID','DEPARTMENT','BATCH YEAR']
                """
                when file template has less columns than the valid columns_list, 
                then data will be saved in mismatched order in database, to avoid that, 
                this validation error is needed."""
                for worksheet in worksheets:
                    excel_data = pd.read_excel(file, worksheet)
                    columns_names = excel_data.columns.values.tolist()

                    fields_list = [
                        column for column in valid_columns_list if column not in columns_names
                    ]
                    if fields_list:
                        fields_dict[worksheet] = fields_list

                    value_list = []
                    for index, row in enumerate(wb[worksheet].iter_rows()):
                        if index == 0:
                            continue

                        row_data = list()
                        for cell in row:
                            row_data.append(cell.value)

                        if not row_data[0]:
                            value_list.append(index)
                        
                    if value_list:
                        value_dict[worksheet] = value_list

                if fields_dict:
                    nl = "\n"
                    messages.error(
                        request,
                        f"Failed upload: mentioned columns should be listed out in respective worksheets :{nl}{nl.join(f'{key}: {value}' for key,value in fields_dict.items())}",
                    )
                    return redirect(
                        reverse_lazy("urlsearch:getProfile")
                    )
                
                if value_dict:
                    nl = "\n"
                    messages.error(
                        request,
                        f"Failed upload: some values are missing in 'LINKEDIN PROFILE URL' column in respective worksheets :{nl}{nl.join(f'{key}: {value}' for key,value in value_dict.items())}",
                    )
                    return redirect(
                        reverse_lazy("urlsearch:getProfile")
                    ) 
                task=get_data_for_xlxs.delay(file,user)
                TakeTaskInfo(task,user,file,task_name)

                # get_data_for_xlxs(file,user)
            else:
                messages.error(
                        request,
                        f"Failed upload: File type does not exist. Please upload text or excel file.",
                    )
                return redirect(
                        reverse_lazy("urlsearch:getProfile")
                    )   
        messages.success(
                    request,
                    f" Upload Successfully: Data has been uploaded successfully. When we complete the task then send the mail you ",
                )
        return redirect(
                    reverse_lazy("urlsearch:getProfile")
                )  
    


def TakeTaskInfo(task,user,file,task_name):
    task_info=TasksInfo.objects.create(task=task.task_id,task_name=task_name,file=file,user=user,college_name=user.college_name)




@method_decorator(login_required, name="dispatch")
class SetUpdateScheduler(UpdateView):
    model = PeriodicTask
    template_name = "schedule.html"
    fields = ['name', 'interval']
    
    def post(self, request, *args, **kwargs):
        task_name = request.POST.get("name")
        time = request.POST.get('time')
        periodic_obj = self.model.objects.filter(name = task_name).first()
        daily_schedule, created = IntervalSchedule.objects.get_or_create(
            every=int(time),
            period=IntervalSchedule.DAYS,
        )
        periodic_obj.interval=daily_schedule
        periodic_obj.save()
        return JsonResponse({"edit_schedule_status" : "success"})

@method_decorator(login_required, name="dispatch")
class SetDeleteeScheduler(DeleteView):
    model = PeriodicTask
    success_url = reverse_lazy("urlsearch:scheduler")
    queryset = PeriodicTask.objects.all() 


@method_decorator(login_required, name="dispatch")
class SetScheduler(TemplateView):
    template_name= "schedule.html"
    success_url = reverse_lazy("urlsearch:scheduler")

    def get_context_data(self, **kwargs):
        college_obj=College.objects.filter(college_name=self.request.user.college_name).first()

        dept=list(BasicInformation.objects.using(self.request.user.college_name.college_name).all().values_list("department", flat=True).distinct())
        print(dept,"---department")
        college_name = college_obj.college_name.lower().replace(" ","_")

        days=list(range(15,500,15))

        tasks_list = list(PeriodicTask.objects.filter(name__contains=college_name))
        avail_batch_years=BasicInformation.objects.using(self.request.user.college_name.college_name).all().values_list("batch_year", flat=True).distinct()
        context =  super().get_context_data(**kwargs, department_list = dept ,record_year = avail_batch_years, days=days, tasks_list = tasks_list )

        return context

    def post(self, request, *args, **kwargs):
        user=request.user
        if request.user.is_role_admin or request.user.is_role_superuser:
            record_year=request.POST.get('record_year')
            department_name=request.POST.get('department_name')
            description=request.POST.get('description')
            user_college_name = user.college_name.college_name
            day=request.POST.get('time')
            temp_task_name=str(department_name+"_"+record_year+"_"+user_college_name)
            task_name=temp_task_name.lower().replace(' ','_')
            obj=PeriodicTask.objects.filter(name=task_name).first()
            if obj:
                messages.error(
                        request,
                        f"This Schedular aready applied",
                    )
                return redirect(
                        reverse_lazy("urlsearch:scheduler")
                    )   
            length=CreateScheduleTask(task_name,description,department_name,record_year,user.email, user_college_name, day, None )
            if length==0:
                messages.error(
                        request,
                        f"You have selected 0 urls",
                    )
                return redirect(
                        reverse_lazy("urlsearch:scheduler")
                    )   
            messages.success(
                    request,
                    f" Your schedular added success fully,when this schedular trigger than we send you mail ",
                )
            return redirect(
                    reverse_lazy("urlsearch:scheduler")
                )  
        

def CreateScheduleTask(task_name,description ,department_name ,record_year,user_email, user_college_name , day, public_list= None):
    daily_schedule, created = IntervalSchedule.objects.get_or_create(
        # every=1,
        # period=IntervalSchedule.MINUTES,
        every=int(day),
        period=IntervalSchedule.DAYS,
    )
    DB=user_college_name

    publicid_list=[]
    if public_list:
        publicid_list = public_list
        if len(publicid_list)==1 and publicid_list[0]=='':
            return 0

    else:
        department_obj=Department.objects.filter(department_name=department_name).first()
        objs=list(BasicInformation.objects.using(DB).filter(department=department_obj.department_name,batch_year=record_year))
        for obj in objs:
            publicid_list.append(obj.PublicId)

    if len(public_list)==0:
        return 0
    # daily_schedule,created=CrontabSchedule.objects.get_or_create(day_of_month='*/'+day)
    else:
        task=PeriodicTask.objects.create(
            interval=daily_schedule,
            name=task_name,
            task="UrlSearch.tasks.update_periodically",
            start_time=timezone.now(),
            description=description+"\n( Total Number of URL : "+str(len(publicid_list))+" )",
            kwargs=json.dumps({"publicid_list":publicid_list, "user_email": user_email,"user_college_name" : user_college_name })
            )
        return len(public_list)
    

import json
from TrackYourAlumni.celery import app
#to rigger instaance task
def trigger_task(request,pk):
    task_id=pk
    print(task_id,"--------")
    obj=PeriodicTask.objects.filter(id=task_id).first()
    print(type(obj.kwargs),"--------")
    result = app.send_task(obj.task,kwargs=json.loads(obj.kwargs))
    return JsonResponse({"success":"done"})
    
