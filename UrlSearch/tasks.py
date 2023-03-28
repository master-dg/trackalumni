from celery import shared_task
from UrlSearch.function import getapi,get_urlprofile,insert_BasicInformation,insert_CollegeDetails,insert_Skills,get_update_profile,insert_CompanyDetails
from django.conf import settings
from django.core.mail import EmailMessage
from celery_progress.backend import ProgressRecorder
from io import BytesIO
import openpyxl
import os
from TrackYourAlumni.settings import USER_NAME1,PASSWORD1,USER_NAME2,PASSWORD2,USER_NAME3,PASSWORD3


def get_api_try():
    try:
        print(USER_NAME1,PASSWORD1,"----try1")
        api1=getapi(email=USER_NAME1,password=PASSWORD1)
        print(api1)
        return api1
    except:
        try:
            print(USER_NAME2,PASSWORD2,"----try2")
            api2=getapi(email=USER_NAME2,password=PASSWORD2)
            print(api2)
            return api2
        except:
            try:
                print(USER_NAME3,PASSWORD3,"----try3")
                api3=getapi(email=USER_NAME3,password=PASSWORD3)
                print(api3)
                return api3
            except:
                print("Now Please Wait")
                return 0




# Authentication
# api= getapi(email="ghelanimeet2002@gmail.com",password="Meetghelani@2002")
# api=get_api_try()

# print(USER_NAME1,type(USER_NAME1),PASSWORD1,type(PASSWORD1),"---1")
# print(USER_NAME2,type(USER_NAME2),PASSWORD2,type(PASSWORD2),"---2")
# print(USER_NAME3,type(USER_NAME3),PASSWORD3,type(PASSWORD3),"---3")

# api=getapi(email=USER_NAME1,password=PASSWORD1)


# fetch people's details by searching from url
@shared_task(bind=True)
def get_data_for_xlxs(self,file,user):
    progress_recorder=ProgressRecorder(self)

    wb = openpyxl.load_workbook(file)
    worksheets = wb.sheetnames   
    # iterating over the rows and
    # getting value from each cell in row            
    for worksheet in worksheets:
        count=wb[worksheet].max_row
        for index, row in enumerate(wb[worksheet].iter_rows()):
            if index == 0:
                continue

            row_data = list()
            for cell in row:
                row_data.append(cell.value)
            print('*'*200)
            DB=user.college_name.college_name
            print(row_data,"-----------row data")
            try:
                details=get_urlprofile(api,row_data[0])
            except:

                api= get_api_try()
                details=get_urlprofile(api,row_data[0])

            public_id=details[0]
            profile=details[1]
            obj=insert_BasicInformation(profile,public_id,DB)
            obj.batch_year=row_data[4]
            obj.department=row_data[3]
            obj.college_id=row_data[2]
            obj.save()
            insert_CollegeDetails(profile,public_id,obj,DB)
            insert_CompanyDetails(profile,public_id,obj,DB)
            insert_Skills(api,public_id,obj,DB)

            
            progress_recorder.set_progress(int(index),count-1)


    send_file_upload_mail(file, user, wb)
    return 'Done '+str(count-1)+' out of '+str(count-1)+"(100%)"


@shared_task(bind=True)
def get_data_for_text(self, file,user,department,record_year):
    progress_recorder=ProgressRecorder(self)
    publicid_list=[]
    index=0
    for line in file.readlines():
        publicid_list.append(str(line,'utf-8'))
    for publicid in enumerate(publicid_list):
        index+=1
        count=len(publicid_list)

        print('*'*200)
        details=get_urlprofile(api,publicid[1])
        public_id=details[0]
        profile=details[1]
        DB=user.college_name.college_name
        obj=insert_BasicInformation(profile,public_id,DB)
        insert_CollegeDetails(profile,public_id,obj,DB) 
        insert_Skills(api,public_id,obj,DB)
        insert_CompanyDetails(profile,public_id,obj,DB)
        obj.batch_year=record_year
        obj.department=department
        obj.save()
        progress_recorder.set_progress(int(index),count)

    send_file_upload_mail(file, user, wb=None)
    return 'Done'

def send_file_upload_mail(file, user, wb = None):
    subject = "Welcome To TrackAlumni! File Uploaded."
    if wb:
        output= BytesIO()
        wb.save(output)
        mail = EmailMessage(subject=subject, body=f"Hello {user.email},\nYou have uploaded attached file successfully for data analysis.", from_email=settings.DEFAULT_FROM_EMAIL, to=[user.email] )
        mail.attach(file.name, output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        mail.send()
    else:
        with file.open('r') as f:
            content = f.read()
        mail = EmailMessage(subject=subject, body=f"Hello {user.email},\nYou have uploaded attached file successfully for data analysis.", from_email=settings.DEFAULT_FROM_EMAIL, to=[user.email] )
        mail.attach(file.name, content , 'text/plain')
        mail.send()
    

@shared_task(bind=True)
def update_periodically(self,**kwargs):
    DB=kwargs.get('user_college_name')
    publicid_list=kwargs.get('publicid_list')
    print(publicid_list,"*"*200)
    
    for publicid in enumerate(publicid_list):
        try:
            details=get_update_profile(api,publicid[1])
            public_id=details[0]
            profile=details[1]
            obj=insert_BasicInformation(profile,public_id,DB)
            insert_Skills(api,public_id,obj,DB)
        except:
            # api= getapi(email="meet32ppsv2020@gmail.com",password="Meet@05102002")
            api=get_api_try()
            details=get_update_profile(api,publicid[1])
            public_id=details[0]
            profile=details[1]
            obj=insert_BasicInformation(profile,public_id,DB)
            insert_Skills(api,public_id,obj,DB)
            
        # public_id=details[0]
        # profile=details[1]
        # obj=insert_BasicInformation(profile,public_id,DB)
        insert_CollegeDetails(profile,public_id,obj,DB) 
        insert_CompanyDetails(profile,public_id,obj,DB)
        # insert_Skills(api,public_id,obj,DB)
    
